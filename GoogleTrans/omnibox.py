#!/usr/bin/env python  
#-*- coding:utf-8 _*-  
'''
@author: jwli9
@contact: jwli9@iflytek.com
'''
import os, codecs, json, time, sys, datetime, pickle, logging, re, random, numpy as np, pandas as pd, matplotlib.pyplot as plt
from tqdm import tqdm
from itertools import groupby
from sklearn.model_selection import train_test_split
from sklearn.metrics import precision_score, recall_score, f1_score, accuracy_score
from pdb import set_trace as stop

def split_train_test(data, test_size):
        return train_test_split(data, test_size=test_size)




def jpath(x, y):
        '''
        :param x: file dir
        :param y: file name
        :return: joint path
        '''
        return os.path.join(x, y)


def jdir(path):
        if not os.path.exists(path):
                os.mkdir(path)
        return path


def db(expression, showValue=True):
        '''
        :param expression: format: 'variable' or "variable"
        :param showValue: bool,whether show variable's content, if False, simply show the type and shape
        :return: None
        '''
        frame = sys._getframe(1)
        print('\n')
        print('-' * 25)
        print('[NAME] : %s' % expression)
        print("[TYPE] :", type(eval(expression, frame.f_globals, frame.f_locals)))
        try:
                print("[SHAPE]:", eval(expression, frame.f_globals, frame.f_locals).shape)
        except:
                pass
        try:
                print("[LEN] :", len(eval(expression, frame.f_globals, frame.f_locals)))
        except:
                pass
        if showValue:
                print('[VALUE]:', repr(eval(expression, frame.f_globals, frame.f_locals)))
        if isinstance(eval(expression, frame.f_globals, frame.f_locals), dict):
                print('[KEYS] :', list(eval(expression, frame.f_globals, frame.f_locals).keys()))
        print('-' * 25)


class omnifile(object):
        def __init__(self, lines=None, size=None, src_path=None, info=None):
                self.lines = lines
                self.size = size
                self.src_path = src_path
                self.info = info

        @classmethod
        def save_json(cls, data, path, encoding='utf-8'):
                with codecs.open(path, 'w', encoding=encoding) as f:
                        json.dump(data, f, ensure_ascii=False)
                        print('Finished dumping json file!')

        @classmethod
        def save_txt(cls, data: list, path, encoding='utf-8'):
                with codecs.open(path, 'w', encoding=encoding) as f:
                        for item in data:
                                f.write(item + '\n')

                print('Finished dumping txt file at {}!'.format(path))

        @classmethod
        def savepkl(cls, data, path):
                with open(path, 'wb') as pklfile:
                        pickle.dump(data, pklfile)
                print('Finished dumping pkl file at {}!'.format(path))


        @classmethod
        def read_(cls, path, split_mark='\t', encoding='utf-8', sheet_id=0, **kwargs):
                if path.endswith('.txt'):
                        new_lines = []
                        size = 0
                        with open(path, 'r', encoding=encoding) as file:
                                lines = file.readlines()
                                size = len(lines)
                                for i in range(size):
                                        line = lines[i].strip().split(split_mark)
                                        new_lines.append(line)

                        return cls(lines=new_lines, size=size, src_path=path)

                elif path.endswith('.json'):
                        import json
                        with open(path, 'r', encoding=encoding) as f:
                                data = json.loads(f.read())
                                size = len(data)
                        return cls(lines=data, size=size, src_path=path)

                elif path.endswith('.xml'):
                        import xmldom
                        domtree = xmldom.parse(path)
                        rootnode = domtree.documentElement
                        instruction = ("XML文件读取后,采用以下函数读取标签:\n"
                                       "对于一个为'paragraph'的标签: nodes_paragraph = rootnode.getElementsByTagName('paragraph')\n"
                                       "之后通过enumerate遍历里面的内容:\n"
                                       "for idx, node in enumerate(nodes_paragraph): ......\n")
                        return cls(lines=rootnode, size=0, src_path=path, info=instruction)

                elif path.endswith('.csv'):
                        f = open(path, 'r', encoding=encoding, errors='ignore')
                        main_content = pd.DataFrame(pd.read_csv(f, header=0, low_memory=False))
                        main_content = main_content.where(main_content.notnull(), 'NAN')
                        info = ("注意:\n"
                                "读取的DataFrame经过预处理,将所有空白项填充了字符串 'NAN' \n"
                                "需要调用哪列,就 data.lines[列名]\n")
                        return cls(lines=main_content, size=len(main_content), src_path=path, info=info)
                elif path.endswith('.xlsx'):
                        import openpyxl
                        # 读取xlsx文件
                        wb = openpyxl.load_workbook(path)
                        ws = wb.worksheets[sheet_id]
                        # sheet页属性：表名、最大行数、最大列数
                        info = 'titles: {}, max_row: {}, max_column: {}'.format(ws.title, ws.max_row, ws.max_column)
                        print(info)
                        columns = ws.columns
                        all_data = []
                        for col in columns:
                                a_col = []
                                for item in col:
                                        a_col.append(item.value)
                                all_data.append(a_col)

                        return cls(lines=all_data, size=len(all_data[0]), info=info)

                elif path.endswith('.pkl'):
                        f = open(path, 'rb')
                        content = pickle.load(f)
                        return cls(lines=content, size=len(content), src_path=path, info=None)

        @classmethod
        def de_nan(cls, df):
                return df.where(df.notnull(), 'NAN')

        @classmethod
        def shape_(cls, path, split_mark='\t', xml_tag=None, encoding='utf-8', **kwargs):
                if path.endswith('.txt'):
                        with open(path, 'r', encoding=encoding) as file:
                                lines = file.readlines()
                                size = len(lines)
                        print('src: {}'.format(path))
                        print('Shape:', size)


                elif path.endswith('.json'):
                        import json
                        with open(path, 'r', encoding=encoding) as f:
                                data = json.loads(f.read())
                                size = len(data)
                        print('src: {}'.format(path))
                        print('Shape:', size)

                elif path.endswith('.xml'):
                        import xmldom
                        domtree = xmldom.parse(path)
                        rootnode = domtree.documentElement
                        nodes_tag = rootnode.getElementsByTagName('{}'.format(xml_tag))
                        size = len(nodes_tag)
                        print('src: {}'.format(path))
                        print('Shape:', size)


                elif path.endswith('.csv'):
                        with open(path, 'r', encoding=encoding, errors='ignore') as f:
                                main_content = pd.DataFrame(pd.read_csv(f, header=0, low_memory=False))
                                print('src: {}'.format(path))
                                print('Shape:', len(main_content))





def inverse_map(my_dict):
        return {v: k for k, v in my_dict.items()}



def find_repeat(source, elmt):  # The source may be a list or string.
    elmt_index = []
    s_index = 0
    e_index = len(source)
    while (s_index < e_index):
        try:
            temp = source.index(elmt, s_index, e_index)
            elmt_index.append(temp)
            s_index = temp + 1
        except ValueError:
            break
    return elmt_index

# [begin]返回连续序列中的连续相同项 ------
def return_group(inputlst):
        groups = [(k, sum(1 for _ in g)) for k, g in groupby(inputlst)]
        cursor = 0
        result = []
        for k, l in groups:
                result.append((k, [cursor, cursor + l - 1]))
                cursor += l
        return result
# [end]返回连续序列中的连续相同项 ------






# [begin]Time Counter ------

class Timer(object):
        def __init__(self):
                self.start_time = None

        @classmethod
        def start_time_recorder(cls):
                start_time = datetime.datetime.now()
                print('Time Counter start ... , now is\n\n {}\n'.format(start_time))
                print('\n')
                print('\n')
                cls.start_time = start_time

        @classmethod
        def pause_code(cls):
                os.system("pause")

        @classmethod
        def stop_time_recorder(cls):
                end_time = datetime.datetime.now()
                time_cost = end_time - cls.start_time
                print('\nAccurate Time Consumption:', time_cost, '\n')
                time_seconds = (time_cost).seconds
                time_cost = int(time_seconds)
                mod_hour = time_cost % 3600
                real_hour = (time_cost - mod_hour) / 3600
                mod_minute = mod_hour % 60
                real_minute = (mod_hour - mod_minute) / 60
                real_second = mod_minute % 60
                print('\n                          >> Time Used <<\n')
                print('&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&')
                print('                |  ', int(real_hour), 'hours', int(real_minute), 'minutes',
                      int(real_second), 'seconds  |                   ')
                print('&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&')

                time_min = time_seconds / 60
                time_hour = time_min / 60
                print('\nDETAILS:\n')
                print('= {} seconds'.format(time_seconds))
                print('\n= {} minutes'.format(np.around(time_min, decimals=5)))
                print('\n= {} hours'.format(np.around(time_hour, decimals=5)))
                print('\n')
                print('\n')



class brief_timer(object):
        def __init__(self):
                self.start_time = None

        @classmethod
        def start_time_recorder(cls):
                start_time = datetime.datetime.now()
                cls.start_time = start_time

        @classmethod
        def set_start_time(cls):
                start_time = datetime.datetime.now()
                return start_time

        @classmethod
        def pause_code(cls):
                os.system("pause")
        @classmethod
        def get_time_cost(cls):
                end_time = datetime.datetime.now()
                time_cost = end_time - cls.start_time
                return time_cost

        @classmethod
        def stop_time_recorder(cls, structure=False,transform=False):
                end_time = datetime.datetime.now()
                time_cost = end_time - cls.start_time
                print('\nAccurate Time Consumption:', time_cost)
                if structure:
                        time_seconds = (time_cost).seconds
                        time_cost = int(time_seconds)
                        mod_hour = time_cost % 3600
                        real_hour = (time_cost - mod_hour) / 3600
                        mod_minute = mod_hour % 60
                        real_minute = (mod_hour - mod_minute) / 60
                        real_second = mod_minute % 60
                        print('Structured Time Consumption:', int(real_hour), 'hours', int(real_minute), 'minutes',
                              int(real_second), 'seconds')
                        if transform:
                                time_min = time_seconds / 60
                                time_hour = time_min / 60
                                print('= {} seconds'.format(time_seconds),' = {} minutes'.format(np.around(time_min, decimals=5)),' = {} hours'.format(np.around(time_hour, decimals=5)))



# [end]Time Counter ------


# [begin].doc 文件转成 .docx ------
def doc2docx(path):
        '''
        :param path: .doc file path
        :return: .docx file path
        '''
        from win32com import client as wc
        w = wc.Dispatch('Word.Application')
        w.Visible = 0
        w.DisplayAlerts = 0
        doc = w.Documents.Open(path)
        newpath = os.path.splitext(path)[0] + '.docx'
        doc.SaveAs(newpath, 12, False, "", True, "", False, False, False, False)
        doc.Close()
        w.Quit()
        os.remove(path)
        return newpath
# [end].doc 文件转成 .docx ------


# [begin]数据标注协议 ------
# 句子相似任务: 可用于句子相似标注, 序列标注
class SentSimLabeling(object):
        def __init__(self):
                self.out = []
                self.out_item = {'ori': "", 'rec': []}
                self.rec_item = {'text': ""}
                self.out_save_path = ""

        def example(self):
                sent = '只是一个样例'
                self.rec_item['text'] = sent
                self.out_item['rec'].append(self.rec_item)
                self.out_item['ori'] = 0
                self.out.append(self.out_item)
                print(self.out)


        def save_file(self):
                omnifile.save_json(self.out, self.out_save_path)

# [end]数据标注协议 ------

# [begin]根据中文标点符号断句 ------
# import re
def sen_split(txt):
    pattern = r',|\.|/|;|\'|`|\[|\]|<|>|\?|:|"|\{|\}|\~|!|@|#|\$|%|\^|&|\(|\)|-|=|\_|\+|，|。|、|；|‘|’|【|】|·|！| |…|（|）'
    return re.split(pattern, txt)
# [end]根据中文标点符号断句 ------



# [begin]打印函数
def nice_print_prf1(scope,test_acc, test_prec, test_recall, test_f1, best_epoch):
        '''
        :param scope: 'train' or 'dev' or 'test'
        :param test_acc: accuracy
        :param test_prec: precision
        :param test_recall: recall
        :param test_f1: f1
        :param best_epoch: epoch
        :return: None
        '''
        print('{}'.format(scope)+'acc: %.4f%%, precision: %.4f%%, recall: %.4f%%, F1: %.4f%%, epoch: %d' % (test_acc, test_prec, test_recall, test_f1, best_epoch))


def percent(x, total):
        return round(x/total, ndigits=4)*100


# [end]打印函数

# [begin]normalization ------
def minmax_norm(lst):
        out=[]
        for x in lst:
            x = float(x - np.min(lst))/(np.max(lst)- np.min(lst))
            out.append(x)
        return out


def z_score_norm(lst):
        out = []
        arr = np.array(lst)
        for x in arr:
                x = float(x - arr.mean()) / arr.std()
                out.append(x)
        return out

def tf_norm(x, keepdims=False, eps=0.0):
        import tensorflow as tf
        return tf.sqrt(tf.reduce_sum(x**2, -1, keepdims=keepdims))+eps

# [end]normalization ------

def computeR2(X, Y):
        import math
        xBar = np.mean(X)
        yBar = np.mean(Y)
        SSR = 0
        varX = 0
        varY = 0
        for i in range(0, len(X)):
                diffXXBar = X[i] - xBar
                diffYYBar = Y[i] - yBar
                SSR += (diffXXBar * diffYYBar)
                varX += diffXXBar ** 2
                varY += diffYYBar ** 2

        SST = math.sqrt(varX * varY)

        return SSR / SST



def list_to_numpy(batch, dtype):
        batch_size = len(batch)
        max_length = max([len(_) for _ in batch])
        output = np.zeros([batch_size, max_length], dtype=dtype)
        for i, sent in enumerate(batch):
                for j, word in enumerate(sent):
                        output[i,j] = word
        return output


def read_vocab_file(vocab_file):
        vocab = {}
        if vocab_file.endswith('.txt'):
                for idx, line in enumerate(open(vocab_file, 'r', encoding='utf-8')):
                        items = line.strip().split('\t')
                        word = items[0]
                        vocab[word] = idx
        if vocab_file.endswith('.pkl'):
                fvocab = omnifile.read_(vocab_file).lines
                for idx, word in enumerate(fvocab):
                        vocab[word] = idx
        return vocab

# [begin]标点符号归一化


# [end]标点符号归一化


# [begin] sequence padding
def pad_sequence(alist, pad_idx, max_len):
        import copy
        output = copy.deepcopy(alist)
        while len(output) < max_len:
                output.extend([pad_idx])
        return output

def tf_pad_seq_v114(input_tensor, seq_len, max_len, pad_id=0):
        import tensorflow as tf
        ranklst = [0,0]*np.rank(input_tensor)
        paddings = tf.constant(ranklst.extend([0, max_len - seq_len]))
        out = tf.pad(input_tensor, paddings, "CONSTANT", constant_values=pad_id)
        return out
# [end] sequence padding


# [begin] tfs

def override_default_initializers(ckpt, variables):
        if not ckpt:
                initialized_var_names = []



# [end] tfs

# [begin] logging
class Logger(object):

        @classmethod
        def create(cls, path, train):
                fname = 'log.train.txt' if train else 'log.test.txt'
                if train:
                        jdir(path)
                else:
                        raise FileNotFoundError(path)
                log_file = jpath(path, fname)
                formatter = logging.Formatter(fmt='%(asctime)s %(message)s', datefmt='%m/%d %H:%M:%S')







# [end] logging

# [begin] plotting
def draw_scatter(label, pred, step):
        label_1 = [label[i] for i in range(0, len(label), step)]
        pred_1 = [pred[i] for i in range(0, len(pred), step)]
        x = range(0, len(label), step)
        plt.figure(figsize=(10, 5.5))
        plt.scatter(x=x, y=label_1, c='r')
        plt.scatter(x=x, y=pred_1, c='g')
        plt.show()

def draw_plot(label, pred, step):
        label_1 = [label[i] for i in range(0, len(label), step)]
        pred_1 = [pred[i] for i in range(0, len(pred), step)]
        x = range(0, len(label), step)
        plt.figure(figsize=(10, 5.5))
        plt.plot(label_1, c='r')
        plt.plot(pred_1, c='g')
        plt.show()

# [end] plotting


# [begin] metrics
class sk_metrics(object):
        def __init__(self, pred, label, type='macro'):
                self.pred = pred
                self.label = label
                self.type = type
                self.__compute_metrics()

        def __compute_metrics(self):
                self.acc = accuracy_score(self.label, self.pred)
                self.p = precision_score(self.label, self.pred, average=self.type)
                self.r = recall_score(self.label, self.pred, average=self.type)
                self.f1 = f1_score(self.label, self.pred, average=self.type)
                self.output = (self.acc, self.p, self.r, self.f1)



# [end] metrics

# [begin] interactive sess
def init_sess():
        import tensorflow as tf
        sess = tf.InteractiveSession()
        return sess

def val_(tensor):
        return tensor.eval()
# [end] tf metrics from scratch


# [begin] 自定义进度条
def prgs(count, total):
        sys.stdout.flush()
        print('\r{} / {}'.format(count, total), end='')
        time.sleep(0.01)

def prgs_example():
        brief_timer.start_time_recorder()
        total = 100
        for i in range(total):
                prgs(i + 1, total)

        brief_timer.stop_time_recorder()


def flush_prgs():
        sys.stdout.flush()
        time.sleep(0.01)

# [end] 自定义进度条


# [start] list切刀
class LstSplitter(object):
        '''
        src: ['我','是','中','国','人','<sep>','我','爱','中','国','<sep>']
        return: [['我','是','中','国','人'],['我','爱','中','国']]
        '''
        def __init__(self, lst, mark=None, type_str=True):
                self.src_lst = lst
                self.mark = mark
                self.out = []
                if type_str:
                        self.string = ''.join(lst)
                        self.splits = self.string.split(self.mark)
                        for item in self.splits:
                                item_lst = [char for char in item]
                                if item_lst:
                                        self.out.append(item_lst)
                else:
                        idx = self.find_repeat(self.src_lst, self.mark)
                        spans = []
                        start=0
                        for id in idx:
                                end = id
                                spans.append((start, end))
                                start = id+1
                        for span in spans:
                                self.out.append(self.src_lst[span[0]:span[1]])

        def find_repeat(self, source, elmt):  # The source may be a list or string.
                elmt_index = []
                s_index = 0
                e_index = len(source)
                while (s_index < e_index):
                        try:
                                temp = source.index(elmt, s_index, e_index)
                                elmt_index.append(temp)
                                s_index = temp + 1
                        except ValueError:
                                break
                return elmt_index

# [end] list切刀






# [start] 分句器
class SentSplitter(object):
        def __init__(self, src_txt, flag='qa'):
                self.src_txt = src_txt
                self.flag = flag
                assert self.flag in ['qa', 'sentence', 'comma'], f"must choose a flag in ['qa', 'sentence', 'comma']"
                if self.flag == 'qa':
                        self.output = self.split_qa()


        def split_qa(self):
                output = []
                idx1 = self.find_repeat(self.src_txt, '问:')
                idx2 = self.find_repeat(self.src_txt, '答:')
                idx = idx1 + idx2
                assert len(idx) % 2 == 0
                idx = sorted(idx)
                for i in range(len(idx)):
                        if i + 1 < len(idx):
                                output.append(self.src_txt[idx[i]:idx[i + 1]])
                        else:
                                output.append(self.src_txt[idx[i]:])
                return output


        def find_repeat(self, source, elmt):  # The source may be a list or string.
                elmt_index = []
                s_index = 0
                e_index = len(source)
                while (s_index < e_index):
                        try:
                                temp = source.index(elmt, s_index, e_index)
                                elmt_index.append(temp)
                                s_index = temp + 1
                        except ValueError:
                                break
                return elmt_index
# [end] 分句器


# [start] list to one-hot
def make_one_hot(label_seq:list, num_class):
        label = np.array(label_seq)
        return (np.arange(num_class) == label[:, None]).astype(np.integer).tolist()

def make_one_hot_v2(label_seq:list):
        from collections import Counter
        num_class = len(dict(Counter(label_seq)).keys())
        label = np.array(label_seq)
        return (np.arange(num_class) == label[:, None]).astype(np.integer)
# [end] list to one-hot


# [start] one-hot to list



# [end] one-hot to list

if __name__ == '__main__':
        a = ['我','是','中','国','人','<sep>','我','爱','中','国','<sep>']
        out = LstSplitter(a, mark='<sep>', type_str=False).out
        print(out)
